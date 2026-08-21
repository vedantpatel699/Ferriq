"""
Ferriq - Furnace Skin TI Predictor
Training script.

Reads the uploaded Excel (Rev 1) with three furnace sheets, mapped to
generic identifiers in the dashboard:
  - L3 sheet -> Heater 3 (atmospheric heater, limited instrumentation)
  - L4 sheet -> Heater 2 (vacuum tower heater, 4 passes x 3 skin TIs)
  - Deep COT sheet -> Heater 1 (atmospheric residue heater, 4 passes x 3-4 skin TIs)

For each furnace and each individual radiant-skin thermocouple, trains THREE
XGBoost quantile regressors (P10, P50, P90) using objective='reg:quantileerror'.
The P50 model gives the central 24h-ahead skin TI prediction; the P10 and P90
models give data-driven lower and upper bounds at 80% nominal coverage. This
replaces the older sigma * sqrt(d) heuristic confidence band with proper
quantile regression intervals.

Per-TC modelling lets the dashboard monitor every thermocouple individually.
Turnaround and upset rows are excluded from training. Decoke events (turnarounds
where skin TI dropped 30+ C, indicating tubes were cleaned) reset the run-day
clock - the model uses days_since_run_start, not days_since_data_start, so
fouling trajectories from different runs are aligned at t=0 instead of averaged
together. The hold-out test set's most recent 90 predictions (P10/P50/P90 vs
actual) are bundled per TC so the dashboard can show empirical model accuracy.

Held-out test = last 20 % of dates per furnace.

Distills the trained trees into a JSON file readable by a pure-JS tree
walker in the dashboard. Generic tag aliases are used everywhere -
no plant tag strings end up in the output.

Outputs:
  predictors/furnace-skin-temp/model.json
  predictors/furnace-skin-temp/training-report.md
"""

from __future__ import annotations

import json
import os
import sys
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
import xgboost as xgb
from sklearn.metrics import mean_absolute_error, r2_score

# -----------------------------------------------------------------------------
# Paths
# -----------------------------------------------------------------------------
HERE = Path(__file__).parent
DATA_DIR = HERE / "data"
EXCEL = DATA_DIR / "Furnaces skin TI example Rev 1.xlsx"
MODEL_JSON = HERE / "model.json"
REPORT_MD = HERE / "training-report.md"

# -----------------------------------------------------------------------------
# Sheet -> furnace mapping
# Each entry describes how to parse the raw sheet into a tidy frame:
#   passes:           how many passes
#   ts_col:           column index of the timestamp
#   header_row:       0-based row index where data starts (the row immediately
#                     after the multi-line header)
#   tag_cols:         dict mapping generic alias -> column index
# -----------------------------------------------------------------------------

# Heater 1 (Deep COT sheet) - daily cadence, ~1900 rows.
# Header layout: row 1 indexes plant area names, row 2 plant tags, row 3 units,
# data starts on row 4 (0-based). Timestamps in col 0.
HEATER_1 = {
    "key": "heater_1",
    "label": "Heater 1 (atmospheric residue, 4-pass)",
    "sheet": "Deep COT",
    "passes": 4,
    "ts_col": 0,
    "data_first_row": 4,  # 0-based; openpyxl row 5
    "cadence_hours": 24,
    "cols": {
        # whole-furnace
        "box_t_c":            52,   # (alias)
        "outlet_avg_t_c":     61,   # (alias)
        "outlet_target_c":    63,   # (alias)
        # per-pass: pass flow (m3/h)
        "flow_pass_1":        10,
        "flow_pass_2":        11,
        "flow_pass_3":        12,
        "flow_pass_4":        13,
        # per-pass: convection skin TI (one TC per pass)
        "conv_skin_pass_1":   29,
        "conv_skin_pass_2":   30,
        "conv_skin_pass_3":   31,
        "conv_skin_pass_4":   32,
        # per-pass: crossover T
        "crossover_pass_1":   33,
        "crossover_pass_2":   34,
        "crossover_pass_3":   35,
        "crossover_pass_4":   36,
        # per-pass: radiant skin TIs (4 per pass except pass 3 has 3)
        "rad_skin_p1_a":      37,
        "rad_skin_p1_b":      38,
        "rad_skin_p1_c":      39,
        "rad_skin_p1_d":      40,
        "rad_skin_p2_a":      41,
        "rad_skin_p2_b":      42,
        "rad_skin_p2_c":      43,
        "rad_skin_p2_d":      44,
        "rad_skin_p3_a":      45,
        "rad_skin_p3_b":      46,
        "rad_skin_p3_c":      47,
        "rad_skin_p4_a":      48,
        "rad_skin_p4_b":      49,
        "rad_skin_p4_c":      50,
        "rad_skin_p4_d":      51,
        # per-pass: outlet T
        "outlet_pass_1":      53,
        "outlet_pass_2":      54,
        "outlet_pass_3":      55,
        "outlet_pass_4":      56,
        # per-pass: IP steam injection
        "steam_pass_1":       25,  # (alias)
        "steam_pass_2":       24,
        "steam_pass_3":       23,
        "steam_pass_4":       22,
    },
    "rad_skin_groups": {
        1: ["rad_skin_p1_a", "rad_skin_p1_b", "rad_skin_p1_c", "rad_skin_p1_d"],
        2: ["rad_skin_p2_a", "rad_skin_p2_b", "rad_skin_p2_c", "rad_skin_p2_d"],
        3: ["rad_skin_p3_a", "rad_skin_p3_b", "rad_skin_p3_c"],
        4: ["rad_skin_p4_a", "rad_skin_p4_b", "rad_skin_p4_c", "rad_skin_p4_d"],
    },
}

# Heater 2 (L4 sheet) - vacuum tower heater, 4 passes, 3 skin TIs per pass.
# Header: row 1-5 establish labels, data starts row 5 (0-based 5, openpyxl row 6).
HEATER_2 = {
    "key": "heater_2",
    "label": "Heater 2 (vacuum tower, 4-pass)",
    "sheet": "L4",
    "passes": 4,
    "ts_col": 0,
    "data_first_row": 5,
    "cadence_hours": 4,
    "cols": {
        # per-pass charge flow
        "flow_pass_1":         2,   # (alias)
        "flow_pass_2":        14,   # (alias)
        "flow_pass_3":        26,   # (alias)
        "flow_pass_4":        38,   # (alias)
        # per-pass IP steam
        "steam_pass_1":        3,   # PQPI-3350 (IP STM)
        "steam_pass_2":       15,
        "steam_pass_3":       27,
        "steam_pass_4":       39,
        # per-pass calculated duty (KW) - free engineered features
        "q_conv_pass_1":      11,
        "q_rad_pass_1":       12,
        "q_conv_pass_2":      23,
        "q_rad_pass_2":       24,
        "q_conv_pass_3":      35,
        "q_rad_pass_3":       36,
        "q_conv_pass_4":      47,
        "q_rad_pass_4":       48,
        # per-pass inlet T and crossover T
        "inlet_pass_1":        8,   # (alias)
        "crossover_pass_1":    9,   # (alias)
        "outlet_pass_1":      10,   # (alias)
        "inlet_pass_2":       20,
        "crossover_pass_2":   21,
        "outlet_pass_2":      22,
        "inlet_pass_3":       32,
        "crossover_pass_3":   33,
        "outlet_pass_3":      34,
        "inlet_pass_4":       44,
        "crossover_pass_4":   45,
        "outlet_pass_4":      46,
        # whole-furnace
        "stack_t_c":          56,   # (alias)
        "fuel_gas_kg_hr":     53,   # (alias)
        "fuel_gas_p_kpa":     54,   # (alias)
        "comb_air_kg_hr":     52,   # (alias)
        # per-pass radiant skin TIs (3 per pass)
        "rad_skin_p1_a":      58,   # (alias)
        "rad_skin_p1_b":      59,   # (alias)
        "rad_skin_p1_c":      60,   # (alias)
        "rad_skin_p2_a":      61,   # (alias)
        "rad_skin_p2_b":      62,
        "rad_skin_p2_c":      63,
        "rad_skin_p3_a":      64,
        "rad_skin_p3_b":      65,
        "rad_skin_p3_c":      66,
        "rad_skin_p4_a":      67,   # (alias)
        "rad_skin_p4_b":      68,   # (alias)
        "rad_skin_p4_c":      69,   # (alias)
    },
    "rad_skin_groups": {
        1: ["rad_skin_p1_a", "rad_skin_p1_b", "rad_skin_p1_c"],
        2: ["rad_skin_p2_a", "rad_skin_p2_b", "rad_skin_p2_c"],
        3: ["rad_skin_p3_a", "rad_skin_p3_b", "rad_skin_p3_c"],
        4: ["rad_skin_p4_a", "rad_skin_p4_b", "rad_skin_p4_c"],
    },
}

# Heater 3 (L3 sheet) - atmospheric heater, treated as 1-pass with 3 skin TIs
# (one convection, two radiant). Limited instrumentation - included so the
# selector covers all three furnaces, but expect lower fidelity.
HEATER_3 = {
    "key": "heater_3",
    "label": "Heater 3 (atmospheric, limited instrumentation)",
    "sheet": "L3",
    "passes": 1,
    "ts_col": 0,
    "data_first_row": 6,   # openpyxl row 7
    "cadence_hours": 4,
    "cols": {
        "feed_t_c":            1,   # (alias)
        "flow_pass_1":         2,   # (alias)
        "inlet_p_kpa":         3,   # PQpI-2106
        "outlet_p_kpa":        4,   # PQpI-2196
        "conv_skin_pass_1":    5,   # (alias)
        "crossover_pass_1":    6,   # (alias)
        "rad_skin_p1_a":       7,   # (alias)
        "rad_skin_p1_b":       8,   # (alias)
        "outlet_pass_1":       9,   # PQtIC-2121
        "stack_t_c":          10,
        "steam_pass_1":       11,
        "box_radiant_t_c":    13,   # (alias)
        "box_conv_t_c":       14,   # (alias)
        "fuel_gas_p_kpa":     15,
        "fuel_gas_kg_hr":     17,
        "comb_air_kg_hr":     18,
    },
    "rad_skin_groups": {
        1: ["rad_skin_p1_a", "rad_skin_p1_b"],
    },
}

FURNACES = [HEATER_1, HEATER_2, HEATER_3]


# -----------------------------------------------------------------------------
# Sheet loader
# -----------------------------------------------------------------------------
def load_sheet(furnace: dict) -> pd.DataFrame:
    """Pull rows out of a sheet using the column map and return a tidy DataFrame.

    Returns columns: timestamp + all aliases declared in `furnace['cols']`.
    """
    wb = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)
    ws = wb[furnace["sheet"]]

    rows_out = []
    aliases = list(furnace["cols"].keys())
    col_idx = [furnace["cols"][a] for a in aliases]
    ts_col = furnace["ts_col"]
    first_row = furnace["data_first_row"]

    # iter_rows yields tuples; we want them indexed
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < first_row:
            continue
        ts = row[ts_col]
        if not isinstance(ts, datetime):
            continue
        rec = {"timestamp": ts}
        for a, c in zip(aliases, col_idx):
            v = row[c] if c < len(row) else None
            try:
                rec[a] = float(v) if v is not None else np.nan
            except (TypeError, ValueError):
                rec[a] = np.nan
        rows_out.append(rec)
    wb.close()

    df = pd.DataFrame(rows_out)
    df = df.sort_values("timestamp").reset_index(drop=True)
    return df


# -----------------------------------------------------------------------------
# Turnaround / upset detection (excluded from training)
# -----------------------------------------------------------------------------
def detect_turnaround(df: pd.DataFrame, furnace: dict,
                      low_thresh: float = 250.0,
                      buffer_days: float = 3.0) -> pd.Series:
    """Mark rows where the whole furnace is in shutdown/startup transition.

    A turnaround row is one where the max skin TI across ALL thermocouples in
    the furnace falls below `low_thresh` (typical: 250 C) or is missing. The
    mask is dilated by `buffer_days` on each side to capture cool-down and
    warm-up periods, which have anomalous trajectories that bias 24h-ahead
    forecasts.
    """
    skin_cols_all = []
    for grp in furnace["rad_skin_groups"].values():
        skin_cols_all.extend(grp)
    skin_cols_all = [c for c in set(skin_cols_all) if c in df.columns]
    if not skin_cols_all:
        return pd.Series([False] * len(df), index=df.index)

    skin_max = df[skin_cols_all].max(axis=1, skipna=True)
    is_down = (skin_max < low_thresh) | skin_max.isna()

    cadence_h = furnace["cadence_hours"]
    buffer_steps = max(1, int(round(buffer_days * 24 / cadence_h)))
    dilated = is_down.copy()
    for _ in range(buffer_steps):
        dilated = (
            dilated
            | dilated.shift(1, fill_value=False)
            | dilated.shift(-1, fill_value=False)
        )
    return dilated


def detect_upset(df: pd.DataFrame, furnace: dict,
                 jump_thresh: float = 50.0) -> pd.Series:
    """Mark rows with abrupt skin TI jumps (sensor glitches or process upsets)."""
    skin_cols_all = []
    for grp in furnace["rad_skin_groups"].values():
        skin_cols_all.extend(grp)
    skin_cols_all = [c for c in set(skin_cols_all) if c in df.columns]
    if not skin_cols_all:
        return pd.Series([False] * len(df), index=df.index)
    skin_max = df[skin_cols_all].max(axis=1, skipna=True)
    delta = skin_max.diff().abs()
    return (delta > jump_thresh).fillna(False)


def detect_decoke_events(df: pd.DataFrame, furnace: dict,
                          turnaround_mask: pd.Series,
                          drop_thresh_c: float = 30.0,
                          window_days: float = 7.0) -> pd.Series:
    """For each turnaround group, decide whether tubes were decoked (cleaned).

    Compare mean max-skin-TI in the `window_days` BEFORE the shutdown vs the
    `window_days` AFTER restart. If post-restart mean is lower by at least
    `drop_thresh_c` degrees, the turnaround included a decoke and the run
    counter resets at the END of that turnaround group.

    Returns: a boolean Series the same length as df, True at the row where
    each detected decoke completes (i.e., the first non-turnaround row after a
    decoke shutdown).
    """
    skin_cols_all = []
    for grp in furnace["rad_skin_groups"].values():
        skin_cols_all.extend(grp)
    skin_cols_all = [c for c in set(skin_cols_all) if c in df.columns]
    if not skin_cols_all:
        return pd.Series([False] * len(df), index=df.index)

    skin_max = df[skin_cols_all].max(axis=1, skipna=True)
    cadence_h = furnace["cadence_hours"]
    window_steps = max(1, int(round(window_days * 24.0 / cadence_h)))

    decoke = pd.Series([False] * len(df), index=df.index)
    in_ta = False
    ta_start = None
    for i in range(len(df)):
        if turnaround_mask.iloc[i] and not in_ta:
            in_ta = True
            ta_start = i
        elif (not turnaround_mask.iloc[i]) and in_ta:
            # turnaround group just ended at row i-1
            ta_end = i - 1
            pre_lo = max(0, ta_start - window_steps)
            pre_hi = ta_start
            post_lo = i
            post_hi = min(len(df), i + window_steps)
            pre_avg = skin_max.iloc[pre_lo:pre_hi].mean()
            post_avg = skin_max.iloc[post_lo:post_hi].mean()
            if pd.notna(pre_avg) and pd.notna(post_avg):
                if post_avg < pre_avg - drop_thresh_c:
                    decoke.iloc[i] = True   # this row begins a fresh run
            in_ta = False
            ta_start = None
    return decoke


def compute_days_since_run_start(df: pd.DataFrame,
                                  decoke_mask: pd.Series) -> pd.Series:
    """For each row, days since the start of the current operating run.

    A new run begins on rows where decoke_mask is True. Before any decoke is
    seen, the run is assumed to start at row 0 (data start). Resets to 0 at
    each decoke event.
    """
    days = np.zeros(len(df))
    run_start_ts = df["timestamp"].iloc[0] if len(df) else None
    for i in range(len(df)):
        if decoke_mask.iloc[i]:
            run_start_ts = df["timestamp"].iloc[i]
        if run_start_ts is None:
            days[i] = 0.0
            continue
        delta = (df["timestamp"].iloc[i] - run_start_ts).total_seconds() / 86400.0
        days[i] = max(0.0, delta)
    return pd.Series(days, index=df.index)


# -----------------------------------------------------------------------------
# Feature engineering
# -----------------------------------------------------------------------------
def engineer_tc(df: pd.DataFrame, furnace: dict, tc_alias: str,
                exclude_mask=None, run_days_series=None):
    """Add target + engineered features for a SINGLE thermocouple.

    Target = this TC's value 24 h ahead.
    Engineered:
        - tc_now                    this TC at current time
        - tc_7d_mean                7-day rolling mean of this TC
        - tc_velocity_c_per_d       (tc_now[t] - tc_now[t-7d]) / 7
        - duty_integral_norm        cumulative pass-flow integral / 1e5
        - days_since_start          proxy for cumulative coke deposition
    Whole-furnace driver columns (box T, fuel gas, etc.) are kept untouched.
    Returns: (DataFrame, pass_num).
    """
    work = df.copy()

    # Sensor-noise mask on every skin TC in the furnace
    skin_cols_all = []
    for grp in furnace["rad_skin_groups"].values():
        skin_cols_all.extend(grp)
    skin_cols_all = [c for c in set(skin_cols_all) if c in work.columns]
    for c in skin_cols_all:
        m_bad = (work[c] < 150) | (work[c] > 700)
        work.loc[m_bad, c] = np.nan

    # Target: this TC at t+24h
    work["tc_now"] = work[tc_alias]
    cadence_h = furnace["cadence_hours"]
    steps_24h = max(1, int(round(24.0 / cadence_h)))
    work["target_tc_24h"] = work[tc_alias].shift(-steps_24h)

    # Rolling features for this TC
    steps_7d = max(1, int(round(7 * 24.0 / cadence_h)))
    work["tc_7d_mean"] = work["tc_now"].rolling(steps_7d, min_periods=1).mean()
    work["tc_velocity_c_per_d"] = (
        work["tc_now"] - work["tc_now"].shift(steps_7d)
    ) / 7.0

    # Identify which pass this TC belongs to (for pass-flow feature)
    pass_num = None
    for p, grp in furnace["rad_skin_groups"].items():
        if tc_alias in grp:
            pass_num = p
            break

    flow_col = f"flow_pass_{pass_num}" if pass_num is not None else None
    if flow_col and flow_col in work.columns:
        f = work[flow_col].ffill().fillna(0.0)
        work["duty_integral_norm"] = (f * cadence_h).cumsum() / 1.0e5
    else:
        work["duty_integral_norm"] = 0.0

    if run_days_series is not None:
        work["days_since_run_start"] = run_days_series.values
    else:
        # fallback: same as before
        work["days_since_run_start"] = (
            (work["timestamp"] - work["timestamp"].iloc[0]).dt.total_seconds() / 86400.0
        )

    # Drop turnaround / upset rows from training: setting target to NaN here
    # is equivalent to filtering them out in build_xy_tc.
    if exclude_mask is not None:
        work.loc[exclude_mask, "target_tc_24h"] = np.nan

    return work, pass_num


def build_xy_tc(df: pd.DataFrame, furnace: dict, pass_num: int):
    """Build (X, y) for a per-TC model. The DataFrame must already have target_tc_24h
    + tc_now + tc_7d_mean + tc_velocity_c_per_d (output of engineer_tc)."""
    feats = []

    # Per-pass driver columns (the pass this TC belongs to)
    for stem in [
        "flow_pass_{p}",
        "steam_pass_{p}",
        "conv_skin_pass_{p}",
        "crossover_pass_{p}",
        "outlet_pass_{p}",
        "inlet_pass_{p}",
        "q_conv_pass_{p}",
        "q_rad_pass_{p}",
    ]:
        col = stem.format(p=pass_num)
        if col in df.columns:
            feats.append(col)

    # Whole-furnace driver columns
    for c in [
        "box_t_c", "outlet_avg_t_c", "outlet_target_c",
        "stack_t_c", "fuel_gas_kg_hr", "fuel_gas_p_kpa", "comb_air_kg_hr",
        "feed_t_c", "inlet_p_kpa", "outlet_p_kpa",
        "box_radiant_t_c", "box_conv_t_c",
    ]:
        if c in df.columns:
            feats.append(c)

    # Engineered (per-TC)
    for c in ["tc_now", "tc_7d_mean", "tc_velocity_c_per_d",
              "duty_integral_norm", "days_since_run_start"]:
        if c in df.columns:
            feats.append(c)

    X = df[feats].astype(float).values
    y = df["target_tc_24h"].astype(float).values
    valid = ~(np.isnan(X).any(axis=1) | np.isnan(y))
    return X[valid], y[valid], feats, valid


# -----------------------------------------------------------------------------
# XGBoost training + JSON distillation
# -----------------------------------------------------------------------------
def train_quantile_set(X, y, feature_names, max_depth=3, n_estimators=80):
    """Fit P10 / P50 / P90 quantile XGBoost regressors on a chronological 80/20 split.

    Returns:
        (m_p10, m_p50, m_p90, metrics_dict)

    metrics_dict includes:
        n_train, n_test
        test_mae_p50_c       median absolute error of P50 on the held-out test
        test_r2_p50          R^2 of P50 on test set
        empirical_80pct_coverage  % of test points falling in [P10, P90] (target ~80)
        median_band_width_c  median of (P90 - P10) on test set
    """
    n = len(X)
    n_train = int(n * 0.8)
    X_tr, X_te = X[:n_train], X[n_train:]
    y_tr, y_te = y[:n_train], y[n_train:]

    common = dict(
        n_estimators=n_estimators,
        max_depth=max_depth,
        learning_rate=0.06,
        subsample=0.85,
        colsample_bytree=0.85,
        tree_method="hist",
        random_state=42,
    )
    m_p10 = xgb.XGBRegressor(objective="reg:quantileerror", quantile_alpha=0.1, **common)
    m_p50 = xgb.XGBRegressor(objective="reg:quantileerror", quantile_alpha=0.5, **common)
    m_p90 = xgb.XGBRegressor(objective="reg:quantileerror", quantile_alpha=0.9, **common)
    m_p10.fit(X_tr, y_tr)
    m_p50.fit(X_tr, y_tr)
    m_p90.fit(X_tr, y_tr)

    if len(X_te) == 0:
        return m_p10, m_p50, m_p90, {
            "n_train": int(n_train),
            "n_test": 0,
            "test_mae_p50_c": float("nan"),
            "test_r2_p50": float("nan"),
            "empirical_80pct_coverage": float("nan"),
            "median_band_width_c": float("nan"),
        }, None

    pred_p10 = m_p10.predict(X_te)
    pred_p50 = m_p50.predict(X_te)
    pred_p90 = m_p90.predict(X_te)

    mae = float(mean_absolute_error(y_te, pred_p50))
    r2 = float(r2_score(y_te, pred_p50)) if len(y_te) > 1 else float("nan")
    in_band = float(np.mean((y_te >= pred_p10) & (y_te <= pred_p90)) * 100.0)
    band_widths = pred_p90 - pred_p10
    med_w = float(np.median(band_widths))

    metrics = {
        "n_train": int(n_train),
        "n_test": int(n - n_train),
        "test_mae_p50_c": round(mae, 3),
        "test_r2_p50": round(r2, 3),
        "empirical_80pct_coverage": round(in_band, 1),
        "median_band_width_c": round(med_w, 2),
    }
    # Hold-out tuple for the dashboard validation view (last <= 90 rows)
    keep = min(90, len(X_te))
    holdout = {
        "actual": [round(float(v), 2) for v in y_te[-keep:]],
        "p10":    [round(float(v), 2) for v in pred_p10[-keep:]],
        "p50":    [round(float(v), 2) for v in pred_p50[-keep:]],
        "p90":    [round(float(v), 2) for v in pred_p90[-keep:]],
    }
    return m_p10, m_p50, m_p90, metrics, holdout


def distill_trees(model: xgb.XGBRegressor, feature_names) -> dict:
    """Convert XGBoost trees to a compact nested JSON for JS evaluation.

    Format per tree:
        node = {"f": feat_idx, "t": threshold, "l": left_node, "r": right_node, "m": missing_dir}
        leaf = {"v": value}
    """
    booster = model.get_booster()
    trees_json = booster.get_dump(dump_format="json")

    def walk(d):
        if "leaf" in d:
            return {"v": float(d["leaf"])}
        feat = d["split"]
        # XGBoost dumps split as feature name like 'f0' or actual name
        if isinstance(feat, str) and feat.startswith("f"):
            try:
                feat_i = int(feat[1:])
            except ValueError:
                feat_i = feature_names.index(feat) if feat in feature_names else 0
        else:
            feat_i = feature_names.index(feat) if feat in feature_names else 0
        thr = float(d["split_condition"])
        # XGBoost children are listed by nodeid; map yes/no
        children = {c["nodeid"]: c for c in d["children"]}
        left = walk(children[d["yes"]])
        right = walk(children[d["no"]])
        # missing direction: 1 = go to yes (left), 0 = go to no (right)
        m = 1 if d.get("missing") == d.get("yes") else 0
        return {"f": feat_i, "t": thr, "l": left, "r": right, "m": m}

    trees = [walk(json.loads(t)) for t in trees_json]
    # XGBoost auto-learns base_score from training data; pull the learned value
    # from the booster's saved config (get_params returns the user-set None).
    cfg = json.loads(booster.save_config())
    learned_base = None
    def find_key(d, key):
        if isinstance(d, dict):
            for k, v in d.items():
                if k == key:
                    yield v
                yield from find_key(v, key)
        elif isinstance(d, list):
            for v in d:
                yield from find_key(v, key)
    for v in find_key(cfg, "base_score"):
        # XGBoost serializes base_score as a string like "[4.71E2]" - parse it
        try:
            if isinstance(v, str):
                s = v.strip().strip("[]")
                learned_base = float(s)
            elif isinstance(v, list):
                learned_base = float(v[0])
            else:
                learned_base = float(v)
            break
        except (TypeError, ValueError):
            continue
    base = learned_base if learned_base is not None else 0.5
    return {
        "base_score": base,
        "trees": trees,
    }


# -----------------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------------
def main():
    if not EXCEL.exists():
        print(f"ERROR: {EXCEL} not found", file=sys.stderr)
        sys.exit(1)

    out = {
        "version": "1.0",
        "trained_at": datetime.now().strftime("%Y-%m-%d"),
        "horizon_hours": 24,
        "alarm_threshold_c": 475.0,
        "advisory_threshold_c": 460.0,
        "furnaces": {},
    }
    report_lines = [
        "# Furnace Skin TI Predictor - Training Report",
        f"Trained: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "",
        "Forecast horizon: 24 hours ahead.",
        "Target: each individual radiant skin thermocouple - 3 quantile models per TC (P10, P50, P90).",
        "Test split: chronological 80/20. Turnaround + upset rows excluded.",
        "",
    ]

    for fcfg in FURNACES:
        print(f"\n=== {fcfg['label']} ===")
        df = load_sheet(fcfg)
        print(f"  rows loaded: {len(df)}")
        print(f"  span: {df['timestamp'].min()} -> {df['timestamp'].max()}")

        f_entry = {
            "label": fcfg["label"],
            "key": fcfg["key"],
            "passes": fcfg["passes"],
            "cadence_hours": fcfg["cadence_hours"],
            "data_span": {
                "start": str(df["timestamp"].min()),
                "end":   str(df["timestamp"].max()),
                "rows":  int(len(df)),
            },
        }
        report_lines.append(f"## {fcfg['label']}")
        report_lines.append(f"- rows: {len(df)}, span: {df['timestamp'].min()} to {df['timestamp'].max()}")
        report_lines.append(f"- cadence: {fcfg['cadence_hours']} h")
        # --- recent history for the dashboard ---------------------------
        # Take last N days of data and emit max-skin-per-pass + key drivers
        df_recent = df.tail(min(540, len(df))).copy()
        # Apply same sensor-noise mask as training
        skin_cols_all = []
        for grp in fcfg["rad_skin_groups"].values():
            skin_cols_all.extend(grp)
        skin_cols_all = list(set(skin_cols_all))
        for c in skin_cols_all:
            if c in df_recent.columns:
                m_bad = (df_recent[c] < 150) | (df_recent[c] > 700)
                df_recent.loc[m_bad, c] = np.nan
        history = []
        for _, row in df_recent.iterrows():
            rec = {"t": str(row["timestamp"])}
            for p in range(1, fcfg["passes"] + 1):
                cols = fcfg["rad_skin_groups"][p]
                vals = [row[c] for c in cols if c in row.index and not pd.isna(row[c])]
                rec[f"skin_max_p{p}"] = round(max(vals), 2) if vals else None
                # individual TC values - dashboard uses these for per-TC monitoring
                for c in cols:
                    if c in row.index and not pd.isna(row[c]):
                        rec[c] = round(float(row[c]), 2)
                fc = f"flow_pass_{p}"
                if fc in row.index and not pd.isna(row[fc]):
                    rec[f"flow_p{p}"] = round(float(row[fc]), 2)
            for c in ["box_t_c", "box_radiant_t_c", "stack_t_c", "outlet_avg_t_c"]:
                if c in row.index and not pd.isna(row[c]):
                    rec[c] = round(float(row[c]), 2)
            history.append(rec)
        f_entry["history"] = history
        # --- current state (latest row) for what-if defaults --------------
        # Forward-fill so trailing NaNs (incomplete final reading) do not blank out
        # the operating state. Then take the last row as "current".
        df_ff = df_recent.ffill()
        last = df_ff.iloc[-1]
        current = {}
        for c in df_ff.columns:
            if c == "timestamp":
                continue
            v = last[c]
            if pd.notna(v):
                try:
                    current[c] = round(float(v), 3)
                except (TypeError, ValueError):
                    pass
        f_entry["current"] = current

        # ===== Detect turnaround / decoke / upset rows + run boundaries =====
        ta_mask = detect_turnaround(df, fcfg)
        upset_mask = detect_upset(df, fcfg)
        exclude_mask = ta_mask | upset_mask
        n_excl = int(exclude_mask.sum())
        n_total = len(df)
        pct = 100.0 * n_excl / max(1, n_total)
        print(f"  excluded {n_excl}/{n_total} rows ({pct:.1f}%) - turnaround + upset")
        report_lines.append(f"- excluded {n_excl} rows ({pct:.1f}%) from training (turnaround + upset)")

        # Decoke detection - which turnarounds reset the fouling clock?
        decoke_mask = detect_decoke_events(df, fcfg, ta_mask)
        decoke_idxs = [i for i in range(len(decoke_mask)) if decoke_mask.iloc[i]]
        decoke_dates = [str(df["timestamp"].iloc[i]) for i in decoke_idxs]
        print(f"  decoke events detected: {len(decoke_idxs)} - {decoke_dates}")
        report_lines.append(f"- decoke events detected: {len(decoke_idxs)} ({', '.join(d[:10] for d in decoke_dates) if decoke_dates else 'none'})")

        # Days since current run started (resets at each decoke)
        run_days = compute_days_since_run_start(df, decoke_mask)
        # Bundle "current run start" info for the dashboard
        if decoke_idxs:
            run_start_ts = df["timestamp"].iloc[decoke_idxs[-1]]
        else:
            run_start_ts = df["timestamp"].iloc[0]
        f_entry["current_run"] = {
            "start_date": str(run_start_ts),
            "days_into_run": round(float(run_days.iloc[-1]), 1),
            "decoke_dates": [d[:10] for d in decoke_dates],
        }

        # ===== Train ONE model per skin thermocouple =====
        f_entry["tc_models"] = {}
        f_entry["tc_to_pass"] = {}
        for p in range(1, fcfg["passes"] + 1):
            tc_aliases = fcfg["rad_skin_groups"].get(p, [])
            for tc in tc_aliases:
                if tc not in df.columns:
                    continue
                d2, pn = engineer_tc(df, fcfg, tc, exclude_mask, run_days_series=run_days)
                X, y, feats, _ = build_xy_tc(d2, fcfg, pn)
                if len(X) < 100:
                    print(f"    {tc}: skipped (only {len(X)} valid rows)")
                    report_lines.append(f"- {tc} (pass {p}): SKIPPED ({len(X)} valid rows)")
                    continue
                m_p10, m_p50, m_p90, metrics, holdout = train_quantile_set(X, y, feats)
                # Capture the hold-out timestamps too. We need them aligned with
                # `holdout` arrays (last 90 test points).
                if holdout is not None and metrics["n_test"] > 0:
                    # df_engineered indices of test rows
                    test_indices = []
                    seen_train = 0
                    seen_total = 0
                    for i in range(len(d2)):
                        row = d2.iloc[i]
                        x_row = []
                        ok = True
                        for nm in feats:
                            v = row[nm]
                            if pd.isna(v): ok = False; break
                            x_row.append(float(v))
                        if not ok: continue
                        if pd.isna(row["target_tc_24h"]): continue
                        if seen_train < metrics["n_train"]:
                            seen_train += 1
                        else:
                            test_indices.append(i)
                        seen_total += 1
                    keep = min(90, len(test_indices))
                    last_idxs = test_indices[-keep:] if keep > 0 else []
                    holdout["timestamps"] = [str(d2.iloc[ix]["timestamp"]) for ix in last_idxs]

                p10_dump = distill_trees(m_p10, feats)
                p50_dump = distill_trees(m_p50, feats)
                p90_dump = distill_trees(m_p90, feats)
                f_entry["tc_models"][tc] = {
                    "feature_names": feats,
                    "pass": p,
                    "p10": {"base_score": p10_dump["base_score"], "trees": p10_dump["trees"]},
                    "p50": {"base_score": p50_dump["base_score"], "trees": p50_dump["trees"]},
                    "p90": {"base_score": p90_dump["base_score"], "trees": p90_dump["trees"]},
                    "metrics": metrics,
                    "holdout": holdout if holdout is not None else None,
                }
                f_entry["tc_to_pass"][tc] = p
                print(f"    {tc} (pass {p}): MAE_P50={metrics['test_mae_p50_c']} "
                      f"cov80={metrics['empirical_80pct_coverage']}% "
                      f"band_med={metrics['median_band_width_c']}")
                report_lines.append(
                    f"- {tc} (pass {p}): train={metrics['n_train']} test={metrics['n_test']} "
                    f"MAE_P50={metrics['test_mae_p50_c']} C  R2_P50={metrics['test_r2_p50']}  "
                    f"empirical_80%_coverage={metrics['empirical_80pct_coverage']}%  "
                    f"median_band_width={metrics['median_band_width_c']} C"
                )

        out["furnaces"][fcfg["key"]] = f_entry
        report_lines.append("")

    # write outputs
    MODEL_JSON.write_text(json.dumps(out, separators=(",", ":")))
    REPORT_MD.write_text("\n".join(report_lines))

    sz = MODEL_JSON.stat().st_size
    print(f"\nWrote {MODEL_JSON} ({sz:,} bytes)")
    print(f"Wrote {REPORT_MD}")


if __name__ == "__main__":
    main()
